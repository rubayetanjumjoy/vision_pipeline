import unittest
import requests
import time
import datetime
import random
import string
from openpyxl import load_workbook,Workbook


def generate_random_string():
    # Generate a random string of length 10
    return ''.join(random.choice(string.ascii_letters) for i in range(10))


class TestJsonToCsv(unittest.TestCase):

    def setUp(self):
        filename='../csv_files/results.xlsx'
        # Set up the test by loading the workbook and worksheet
        try:
            self.workbook = load_workbook(filename)

            # Select the active worksheet
            self.worksheet = self.workbook.active

        except FileNotFoundError:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet['A1'] = 'device_id'
            self.worksheet['B1'] = 'client_id'
            self.worksheet['C1'] = 'created_at'
            self.worksheet['D1'] = 'license_id'
            self.worksheet['E1'] = 'image_frame'
            self.worksheet['F1'] = 'prob'
            self.worksheet['G1'] = 'tags'
            self.workbook.save(filename)
        self.filename = '../csv_files/results.xlsx'
        self.workbook = load_workbook(self.filename)
        self.worksheet = self.workbook.active

    def test_request(self):
        num_messages = 1000
        preds_per_message = 2
        row_initial = self.worksheet.max_row

        for i in range(num_messages):
            # Send a POST request to the endpoint
            url = 'http://localhost:8000/api/predictions/'

            # image base64 string is fixed and all other field are randomly generated
            data = {
                "device_id": generate_random_string(),
                "client_id": generate_random_string(),
                "created_at": f"{datetime.datetime.now()}",
                "data": {
                    "license_id": generate_random_string(),
                    "preds": [
                        
                        {   
                            "image_frame": "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII=",
                            "prob": round(random.randint(0, 4), 2),
                            "tags": []
                        },
                        {
                            "image_frame": "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII=",
                            "prob": round(random.randint(0, 4), 2),
                            "tags": []
                        }
                    ]
                }
            }

            response = requests.post(url, json=data)
            # checking response is 200 OK
            self.assertEqual(response.status_code, 200)

            time.sleep(1)

        self.setUp()

        row_after_test = self.worksheet.max_row

        new_row = row_after_test - row_initial
        # checking the new rows in the csv file are equal to nums_messages * preds_per_message
        self.assertEqual(new_row, num_messages * preds_per_message)


if __name__ == '__main__':
    unittest.main()
