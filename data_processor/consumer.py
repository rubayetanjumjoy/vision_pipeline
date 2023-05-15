from openpyxl import Workbook, load_workbook

import pika
import json
params=pika.ConnectionParameters('host.docker.internal',heartbeat=60)
connection = pika.BlockingConnection(params)


channel=connection.channel()
channel.queue_declare(queue='json.to.csv')

def write_to_csv(data):
    
    filename='/app/csv_files/results.xlsx'
    # check file exist otherwise create new file
    try:
        workbook = load_workbook(filename)

        # Select the active worksheet
        worksheet = workbook.active

    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'device_id'
        worksheet['B1'] = 'client_id'
        worksheet['C1'] = 'created_at'
        worksheet['D1'] = 'license_id'
        worksheet['E1'] = 'image_frame'
        worksheet['F1'] = 'prob'
        worksheet['G1'] = 'tags'
        workbook.save(filename)

    # converting json object to list object
    rows = []
    preds = data["data"]["preds"]
    for pred in preds:
        if pred["tags"]==[]:
            pred["tags"]=""
        else:
            pred["tags"]="low prob"
            
        
        row = [
        data["device_id"],
        data["client_id"],
        data["created_at"],
        data["data"]["license_id"],
        pred["image_frame"],
        pred["prob"],
        pred["tags"]
        
       ]
        
        worksheet.append(row)
    
    print('Data successfully written to Excel file')
    # Save the workbook
    workbook.save(filename)

def callback(ch,method,properties,body):
    data=json.loads(body)
    
    
    write_to_csv(data)
    



channel.basic_consume(queue='json.to.csv',on_message_callback=callback,auto_ack=True)

channel.start_consuming()
print('Started consuming messages from RabbitMQ')
# channel.close()